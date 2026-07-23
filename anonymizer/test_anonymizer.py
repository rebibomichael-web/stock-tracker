#!/usr/bin/env python3
"""
End-to-end acceptance test for record_anonymizer v2 — synthetic data only.

Exercises the §9 acceptance criteria against the generated fixture:
  (a) anonymized output contains zero real values and zero pattern-class hits
      (checked here independently of the tool's own verifier);
  (b) harvest floor met (and hard-FAILs on shortfall — negative controls);
  (c) page count preserved; the downstream transcript-to-excel parser yields
      identical course/exam/cumulative data under aliases;
  (d) restore round-trip on txt/xlsx exact with zero residual aliases;
  (e) re-run determinism: identical mapping.

Run:  python3 test_anonymizer.py        (needs pymupdf, pdfplumber, openpyxl)
"""

import csv
import datetime
import json
import os
import re
import shutil
import sys
import tempfile

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
import record_anonymizer as ra          # noqa: E402
import make_fixture as mf               # noqa: E402

FAILED = []


def check(cond, label):
    print("  %s  %s" % ("PASS" if cond else "FAIL", label))
    if not cond:
        FAILED.append(label)


# ---------------------------------------------------------------------------
# Downstream-parser snapshot (compatibility oracle)
# ---------------------------------------------------------------------------
# Vendored from the transcript-to-excel skill's scripts/parse_transcripts.py
# (snapshot 2026-07-23, parsing functions only). The skill copy is canonical;
# this copy exists so the acceptance test runs without the skill installed.

def parse_course_line(line):
    line = re.sub(r'([\d.]+/[\d.]+).*$', r'\1', line).strip()
    m = re.match(r'^(\d{4})\s*/\s*(\d+)\s+(\S+)\s+(\S+\*{0,2})\s+(.+)$', line)
    if not m:
        return None
    year, term, school, course_code, rest = m.groups()
    m0 = re.match(r'^(.+?)\s+(CR\*|P\*|F\*|NS)\s+([\d.]+/[\d.]+)$', rest)
    if m0:
        return dict(year=year, term=term, school=school, course_code=course_code,
                    course_name=m0.group(1).strip(), mark=m0.group(2), teacher='',
                    credits=m0.group(3))
    m2 = re.match(r'^(.+?)\s+([A-Za-z0-9]+[*]*)\s+([A-Za-z]\S*(?:\s+[A-Z])?)\s+([\d.]+/[\d.]+)$', rest)
    if m2:
        mark = m2.group(2)
        if re.match(r'^\d+[*]?$|^[A-Z]{1,2}[*]?$', mark):
            return dict(year=year, term=term, school=school, course_code=course_code,
                        course_name=m2.group(1).strip(), mark=mark,
                        teacher=m2.group(3).strip(), credits=m2.group(4))
    m3 = re.match(r'^(.+?)\s+([A-Za-z0-9]+[*]*)\s+([\d.]+/[\d.]+)$', rest)
    if m3:
        mark = m3.group(2)
        if re.match(r'^\d+[*]?$|^[A-Z]{1,2}[*]?$', mark):
            return dict(year=year, term=term, school=school, course_code=course_code,
                        course_name=m3.group(1).strip(), mark=mark, teacher='',
                        credits=m3.group(3))
    return None


def parse_student_page(text):
    lines = text.split('\n')
    student = {'name': '', 'student_id': '', 'address': '', 'phone': '',
               'ofcl': '', 'grade_level': '', 'status': '', 'dob': '',
               'admit_date': '', 'parent': '', 'counselor': '',
               'cumulative_credits_actual': '', 'cumulative_credits_earned': '',
               'cumulative_average': '', 'courses': [], 'exams': []}
    for line in lines:
        m = re.search(r'Name / ID\s*:\s*(.+?)\s*/\s*(\d+)', line)
        if m:
            student['name'], student['student_id'] = m.group(1).strip(), m.group(2)
        m = re.search(r'Address\s*:\s*(.+)', line)
        if m:
            student['address'] = m.group(1).strip()
        m = re.search(r'Ph#\s*:\s*(\S+)', line)
        if m:
            student['phone'] = m.group(1)
        m = re.search(r'Ofcl\s*:\s*(\S+)', line)
        if m:
            student['ofcl'] = m.group(1)
        m = re.search(r'Grade Level\s*:\s*(\S+)', line)
        if m:
            student['grade_level'] = m.group(1)
        m = re.search(r'Status\s*:\s*(\S+)', line)
        if m:
            student['status'] = m.group(1)
        m = re.search(r'DOB\s*:\s*(\S+)', line)
        if m:
            student['dob'] = m.group(1)
        m = re.search(r'Admit Date\s*:\s*(\S+)', line)
        if m:
            student['admit_date'] = m.group(1)
        m = re.search(r'Parent\s*:\s*(.+?)(?:\s+Counselor|$)', line)
        if m:
            student['parent'] = m.group(1).strip()
        m = re.search(r'Counselor\s*:\s*(.+)', line)
        if m:
            student['counselor'] = m.group(1).strip()
        m = re.search(r'Cumulative\s*:\s*Actual Credits / Credits Earned\s+([\d.]+)\s*/\s*([\d.]+)', line)
        if m:
            student['cumulative_credits_actual'] = m.group(1)
            student['cumulative_credits_earned'] = m.group(2)
        m = re.search(r'Cumulative Average\s*:\s*([\d.]+%)', line)
        if m:
            student['cumulative_average'] = m.group(1)
        for frag in re.findall(r'\d{4}\s+Term\s+\d+\s+\S+\s+.+?\s+\d+', line):
            em = re.match(r'^(\d{4})\s+Term\s+(\d+)\s+(\S+)\s+(.+?)\s+(\d+)$', frag.strip())
            if em:
                student['exams'].append({'year': em.group(1), 'term': em.group(2),
                                         'exam_code': em.group(3),
                                         'exam_name': em.group(4).strip(),
                                         'score': em.group(5)})
        for fragment in re.findall(r'\d{4}\s*/\s*\d+\s+\S+\s+\S+\*{0,2}\s+.+?[\d.]+/[\d.]+', line):
            course = parse_course_line(fragment)
            if course:
                student['courses'].append(course)
    return student if student['name'] else None


def parse_transcripts(pdf_path):
    import pdfplumber
    students = []
    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            text = page.extract_text()
            if text and 'Student Permanent Record' in text:
                s = parse_student_page(text)
                if s:
                    students.append(s)
    return students


# ---------------------------------------------------------------------------
# Test phases
# ---------------------------------------------------------------------------

def anon_full_text(path):
    import fitz
    doc = fitz.open(path)
    text = "\n".join(p.get_text() for p in doc)
    doc.close()
    return text


def test_pipeline(tmp):
    fixture = os.path.join(tmp, "fixture_records.pdf")
    mapping_path = os.path.join(tmp, "mapping.json")
    pages, n_students = mf.build_fixture(fixture)
    print("fixture: %d pages, %d students" % (pages, n_students))

    print("\n[1] anonymize + built-in verification")
    ok, out, report = ra.anonymize_pdf(fixture, mapping_path=mapping_path,
                                       expect_students=n_students)
    check(ok, "pipeline reports PASS")
    check(out and out.endswith("_ANON.pdf") and os.path.exists(out),
          "ANON output written")
    if not ok:
        print("\n".join(report))
        return None, None, None
    text = anon_full_text(out)

    print("\n[2] independent residual scan (criterion a)")
    leaked = []
    for s in mf.STUDENTS:
        last, first, osis, dob, phone, address, parent, counselor, _ = s
        for v in ["%s, %s" % (last, first), osis, dob, phone]:
            if v in text:
                leaked.append(v)
        for v in (address, parent, counselor):
            if v and re.search(r"\s+".join(re.escape(t) for t in v.split()), text):
                leaked.append(v)
    for t in mf.TEACHERS:
        if re.search(r"\s+".join(re.escape(x) for x in t.split()), text):
            leaked.append("teacher " + t)
    check(not leaked, "no real values in ANON text (leaked: %r)" % leaked[:3])
    check("MARLOWE" not in text, "shared teacher/student surname fully gone")
    check("BROOKLYN" in text, "borough word survives in school name")
    check("VARGA-QUINN" not in text, "borough-named student gone from header")
    check("11218" not in text and "OCEAN PKWY" not in text, "addresses gone")
    check("Admit Date : 09/07/2022" in text, "non-PII admit date preserved")
    check("Ofcl : 3C3" in text, "non-PII Ofcl preserved")

    print("\n[3] downstream transcript-to-excel parser (criterion c)")
    orig_students = parse_transcripts(fixture)
    anon_students = parse_transcripts(out)
    check(len(orig_students) == len(anon_students) == pages,
          "downstream sees all %d record pages in both files" % pages)
    with open(mapping_path) as fh:
        mapping = json.load(fh)
    by_osis = {osis: rec for osis, rec in mapping["students"].items()}
    tmap = mapping["teachers"]
    mismatches = []
    for o, a in zip(orig_students, anon_students):
        rec = by_osis.get(o["student_id"])
        al = ra.student_aliases(rec["idx"]) if rec else None
        if not rec or a["name"] != al["name"] or a["student_id"] != al["id"]:
            mismatches.append("identity %s" % a["name"])
            continue
        for f in ("ofcl", "grade_level", "status", "admit_date",
                  "cumulative_credits_actual", "cumulative_credits_earned",
                  "cumulative_average"):
            if o[f] != a[f]:
                mismatches.append("%s: %r vs %r" % (f, o[f], a[f]))
        if a["dob"] != al["dob"] or (o["phone"] and a["phone"] != al["phone"]):
            mismatches.append("dob/phone alias for %s" % al["name"])
        if o["exams"] != a["exams"]:
            mismatches.append("exams for %s" % al["name"])
        if len(o["courses"]) != len(a["courses"]):
            mismatches.append("course count %d vs %d for %s"
                              % (len(o["courses"]), len(a["courses"]), al["name"]))
            continue
        # On modifier-mark rows ('B+', '92**') the downstream parser leaks
        # the teacher surname into course_name in BOTH files; normalize
        # surnames to their alias bases so those rows compare equal too.
        surname_map = {}
        for tname, talias in tmap.items():
            base = talias.split()[0]
            surname = tname.rsplit(None, 1)[0] if " " in talias else tname
            surname_map[surname] = base

        def norm(name):
            for s, b in surname_map.items():
                name = name.replace(s, b)
            return name

        for oc, ac in zip(o["courses"], a["courses"]):
            for f in ("year", "term", "school", "course_code", "credits"):
                if oc[f] != ac[f]:
                    mismatches.append("course %s field %s" % (oc["course_code"], f))
            if oc["mark"] != ac["mark"]:
                mismatches.append("mark %r vs %r in %s" % (oc["mark"], ac["mark"],
                                                           oc["course_code"]))
            if norm(oc["course_name"]) != ac["course_name"]:
                mismatches.append("course name %r vs %r" % (oc["course_name"],
                                                            ac["course_name"]))
            if oc["teacher"] and tmap.get(oc["teacher"]) != ac["teacher"]:
                mismatches.append("teacher %r -> %r (expected %r)"
                                  % (oc["teacher"], ac["teacher"],
                                     tmap.get(oc["teacher"])))
    check(not mismatches, "downstream data identical under aliases "
          "(diffs: %r)" % mismatches[:4])
    return out, mapping_path, text


def test_restore(tmp, mapping_path):
    print("\n[4] restore round-trips (criterion d)")
    m = ra.Mapping(mapping_path)
    pairs = m.reverse_pairs()
    al1 = ra.student_aliases(1)
    rec1 = next(r for r in m.data["students"].values() if r["idx"] == 1)

    txt = os.path.join(tmp, "analysis.txt")
    body = ("Summary: %s (%s) is on track. %s needs follow-up.\n"
            "Teacher TCH001 average 85. Counselor C01CNSL notified.\n"
            % (al1["name"], al1["id"], al1["first"]))
    with open(txt, "w") as fh:
        fh.write(body)
    out, issues = ra.restore_file(txt, mapping_path)
    check(not issues, "txt: clean restore reports no issues")
    restored = open(out).read()
    check(rec1["name"] in restored and "212000001" not in body,
          "txt: real name restored")
    check(not re.search(r"\b[LF]\d{3}XX\b|TCH\d{3}|C\d{2}CNSL", restored),
          "txt: zero residual aliases")
    expected = body
    for a, r in pairs:
        expected = expected.replace(a, r)
    check(restored == expected, "txt: exact dictionary round-trip")

    import openpyxl
    xp = os.path.join(tmp, "analysis.xlsx")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Student Name", "ID", "Phone", "DOB", "Note"])
    ws.append([al1["name"], int(al1["id"]), int(al1["phone"]),
              datetime.datetime.strptime(al1["dob"], "%m/%d/%Y"),
              "flag %s for honors" % al1["first"]])
    al8 = ra.student_aliases(8)
    ws.append([al8["name"], al8["id"], al8["phone"], al8["dob"], ""])
    wb.save(xp)
    out, issues = ra.restore_file(xp, mapping_path)
    check(not issues, "xlsx: clean restore reports no issues")
    wb2 = openpyxl.load_workbook(out)
    cells = [c.value for row in wb2.active.iter_rows() for c in row if c.value]
    joined = " | ".join(str(c) for c in cells)
    check(rec1["name"] in joined and "MARLOWE, FENN" in joined,
          "xlsx: names restored cell-by-cell")
    check("XX" not in joined and "TCH" not in joined,
          "xlsx: zero residual aliases")
    check(str(rec1["real"]["phone"]) in joined,
          "xlsx: numeric-typed ID/phone cells restored (not just string cells)")
    check(rec1["real"]["dob"] in joined,
          "xlsx: date-typed DOB cell restored")

    csvp = os.path.join(tmp, "analysis.csv")
    with open(csvp, "w", newline="") as fh:
        csv.writer(fh).writerow(
            [al1["name"], al1["id"], al1["parent"].split(", ")[1], "on track"])
    out, issues = ra.restore_file(csvp, mapping_path)
    check(not issues, "csv: clean restore reports no issues")
    with open(out, newline="") as fh:
        rows = list(csv.reader(fh))
    check(len(rows) == 1 and len(rows[0]) == 4,
          "csv: restoring a comma-bearing value does not shift columns "
          "(row stays 4 fields, %r)" % (rows[0] if rows else None))
    check(rows[0][0] == rec1["name"] if rows else False,
          "csv: name cell restored exactly")


def test_pdf_restore(tmp, anon_path):
    print("\n[5] pdf restore (spot checks)")
    out, issues = ra.restore_file(anon_path, os.path.join(tmp, "mapping.json"))
    # The full fixture deliberately includes names too long to redraw into
    # their (much shorter) alias's redacted box — a genuine physical PDF
    # constraint, not a bug. The FIX under test is that this is now
    # SURFACED as an issue (a caller can't miss it) instead of silently
    # leaving that student's name blank with a clean-looking return value.
    check(bool(issues) and "could not be redrawn" in issues[0],
          "pdf: a too-long-to-refit real name is surfaced as an issue, "
          "not silently dropped (issues: %r)" % issues)
    text = anon_full_text(out)
    check("AARDWOLF, QUIMBY" in text, "pdf: student name restored")
    check("212000001" in text, "pdf: OSIS restored")
    check("MARLOWE" in text, "pdf: teacher restored")

    # A restore with names short enough to fit must report zero issues —
    # the surfaced-issue path above must not fire unconditionally.
    import fitz
    doc = fitz.open()
    page = doc.new_page()
    m = ra.Mapping(os.path.join(tmp, "mapping.json"))
    a1 = ra.student_aliases(1)
    page.insert_text((50, 100), "%s / %s" % (a1["name"], a1["id"]), fontsize=10)
    shortp = os.path.join(tmp, "short.pdf")
    doc.save(shortp)
    doc.close()
    out3, issues3 = ra.restore_file(shortp, os.path.join(tmp, "mapping.json"))
    check(not issues3, "pdf: a short, fitting name restores with zero issues")
    check("AARDWOLF, QUIMBY" in anon_full_text(out3),
          "pdf: short name actually restored")

    # Adjacent-line restore must not skip a hit merely because its rect
    # overlaps a previously-claimed one at normal line spacing.
    doc = fitz.open()
    page = doc.new_page()
    a2 = ra.student_aliases(2)
    page.insert_text((50, 100), "%s / %s" % (a1["name"], a1["id"]), fontsize=10)
    page.insert_text((50, 112), "%s / %s" % (a2["name"], a2["id"]), fontsize=10)
    adjp = os.path.join(tmp, "adjacent.pdf")
    doc.save(adjp)
    doc.close()
    out2, issues2 = ra.restore_file(adjp, os.path.join(tmp, "mapping.json"))
    text2 = anon_full_text(out2)
    check("AARDWOLF, QUIMBY" in text2 and "VARGA-QUINN, BROOKLYN" in text2,
          "pdf: BOTH adjacent-line aliases restored (containment, not "
          "any-overlap, gates the substring skip)")


def test_determinism(tmp):
    print("\n[6] re-run determinism (criterion e)")
    fixture = os.path.join(tmp, "fixture_records.pdf")
    mp1 = os.path.join(tmp, "mapping.json")
    with open(mp1) as fh:
        m1 = json.load(fh)
    # re-run against the SAME mapping
    os.remove(fixture.replace(".pdf", "_ANON.pdf"))
    ok, _, _ = ra.anonymize_pdf(fixture, mapping_path=mp1)
    with open(mp1) as fh:
        m2 = json.load(fh)
    for m in (m1, m2):
        m.pop("created", None), m.pop("updated", None)
    check(ok and m1 == m2, "same mapping after re-run")
    # from-scratch regeneration assigns identical aliases (same page order)
    mp3 = os.path.join(tmp, "mapping_fresh.json")
    os.remove(fixture.replace(".pdf", "_ANON.pdf"))
    ok, _, _ = ra.anonymize_pdf(fixture, mapping_path=mp3)
    with open(mp3) as fh:
        m3 = json.load(fh)
    m3.pop("created", None), m3.pop("updated", None)
    check(ok and m1 == m3, "identical mapping from a fresh start")


def test_fail_paths(tmp):
    print("\n[7] hard-FAIL paths (criterion b + safety)")
    import fitz
    # (i) a Name / ID label whose value cannot be harvested
    bad = os.path.join(tmp, "bad.pdf")
    doc = fitz.open()
    page = doc.new_page(width=612, height=792)
    page.insert_text((30, 42), "Student Permanent Record", fontname="cour", fontsize=7)
    page.insert_text((14, 62), "Name / ID  :  GLITCHED RECORD NO ID", fontname="cour", fontsize=7)
    doc.save(bad)
    doc.close()
    ok, out, report = ra.anonymize_pdf(bad, mapping_path=os.path.join(tmp, "m_bad.json"))
    check(not ok, "unharvestable Name / ID page FAILs")
    check(not os.path.exists(os.path.join(tmp, "bad_ANON.pdf")),
          "no uploadable output written on floor FAIL")

    # (ii) --expect mismatch FAILs
    fixture = os.path.join(tmp, "fixture_records.pdf")
    ok, _, _ = ra.anonymize_pdf(fixture, mapping_path=os.path.join(tmp, "m_e.json"),
                                expect_students=99)
    check(not ok, "--expect mismatch FAILs")

    # (iii) a rogue out-of-header name occurrence must be caught by the
    # residual scan and the output quarantined as _ANON_FAILED
    rogue = os.path.join(tmp, "rogue.pdf")
    doc = fitz.open()
    page = doc.new_page(width=612, height=792)
    page.insert_text((30, 42), "Student Permanent Record", fontname="cour", fontsize=7)
    page.insert_text((14, 62), "Name / ID  :  ROGUEBERG, ZINNIA  /  212999999",
                     fontname="cour", fontsize=7)
    page.insert_text((14, 200), "note for ROGUEBERG, ZINNIA in free text",
                     fontname="cour", fontsize=7)
    doc.save(rogue)
    doc.close()
    ok, out, report = ra.anonymize_pdf(rogue, mapping_path=os.path.join(tmp, "m_r.json"))
    check(not ok, "rogue extra-header occurrence FAILs the run")
    check(out and out.endswith("_ANON_FAILED.pdf") and os.path.exists(out),
          "output quarantined with _ANON_FAILED suffix")
    joined_report = "\n".join(report)
    check("ROGUEBERG" not in joined_report and "ZINNIA" not in joined_report,
          "FAIL report masks the real value instead of quoting it verbatim "
          "(reports are what a user pastes when asking for help)")

    # (iv) a real value stashed in PDF metadata (Title/Author) must FAIL —
    # a redaction pass that only touches page text would ship it untouched
    meta = os.path.join(tmp, "meta.pdf")
    doc = fitz.open()
    page = doc.new_page(width=612, height=792)
    page.insert_text((30, 42), "Student Permanent Record", fontname="cour", fontsize=7)
    page.insert_text((14, 62), "Name / ID  :  ROGUEBERG, ZINNIA  /  212999998",
                     fontname="cour", fontsize=7)
    doc.set_metadata({"title": "Permanent Record - ROGUEBERG, ZINNIA"})
    doc.save(meta)
    doc.close()
    ok, out, report = ra.anonymize_pdf(meta, mapping_path=os.path.join(tmp, "m_meta.json"))
    check(not ok, "real value in PDF metadata (Title) FAILs the run")

    # (v) a case-variant of a known real value FAILs, not just warns
    case = os.path.join(tmp, "case.pdf")
    doc = fitz.open()
    page = doc.new_page(width=612, height=792)
    page.insert_text((30, 42), "Student Permanent Record", fontname="cour", fontsize=7)
    page.insert_text((14, 62), "Name / ID  :  ROGUEBERG, ZINNIA  /  212999997",
                     fontname="cour", fontsize=7)
    page.insert_text((14, 200), "see Rogueberg, Zinnia re: schedule",
                     fontname="cour", fontsize=7)
    doc.save(case)
    doc.close()
    ok, out, report = ra.anonymize_pdf(case, mapping_path=os.path.join(tmp, "m_case.json"))
    check(not ok, "case-variant of a real value FAILs (not just warns)")


def main():
    tmp = tempfile.mkdtemp(prefix="anon_test_")
    try:
        anon_path, mapping_path, _ = test_pipeline(tmp)
        if anon_path:
            test_restore(tmp, mapping_path)
            test_pdf_restore(tmp, anon_path)
            test_determinism(tmp)
        test_fail_paths(tmp)
    finally:
        shutil.rmtree(tmp, ignore_errors=True)
    print("\n%s — %d failure(s)" % ("ALL PASS" if not FAILED else "FAILED", len(FAILED)))
    for f in FAILED:
        print("  ! %s" % f)
    return 1 if FAILED else 0


if __name__ == "__main__":
    sys.exit(main())
