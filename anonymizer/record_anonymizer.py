#!/usr/bin/env python3
"""
NYC DOE permanent-record anonymizer — v2.

Pseudonymizes student permanent-record PDFs (names, OSIS IDs, DOBs, phones,
addresses, parents, counselors, teachers) before they are uploaded anywhere,
and restores real values into returned outputs (txt/csv/xlsx/pdf). Runs
entirely on the local machine; real PII never leaves it.

v2 replaces v1's line-regex harvest (which silently found 0/133 students
because PyMuPDF's label whitespace differs from the text the regexes were
validated on) with a word-coordinate harvest: labels and values are located
as word boxes via page.get_text("words"), so detection and redaction share
one coordinate space and no line reconstruction is involved.

Safety posture (all hard requirements, see anonymizer/README.md):
  * Harvest floor — every page bearing a "Name / ID" label must yield a
    (name, OSIS) pair or the run FAILs; a vacuous pass is impossible.
  * Field harvesting is label-bounded, not distance-bounded: a detached or
    wrapped value is searched for by walking forward from its label until
    the next label, the course-table anchor, or the header column ends —
    never by a fixed point tolerance that can silently undershoot.
  * Alias-presence check covers every harvested PII field, not just name/ID:
    if a redacted value's alias could not be drawn, the run FAILs.
  * Residual scan — the anonymized output is re-extracted (page text +
    metadata) and scanned for every harvested real value AND for pattern
    classes (LAST, FIRST shapes incl. short surnames/single-initial given
    names, Title-Case "First Last" shapes, street addresses, DOB-range
    dates, 9/10-digit numbers outside the synthetic ranges). Class hits
    and case-variants of known values both FAIL the run.
  * Non-PII header fields (Ofcl, Admit/Discharge/Graduation Date, Grade
    Level, Status, Cumulative Average) are diffed original-vs-anonymized
    and must match exactly, so a redaction that clips a neighboring field
    is caught, not just course/exam data.
  * Any exception (including Ctrl-C) during redaction/verification leaves
    no file under the clean "_ANON.pdf" name — output is written under a
    ".partial" name and only renamed to "_ANON.pdf" after a full PASS, or
    to "_ANON_FAILED.pdf" otherwise.

The mapping file (anonymizer_mapping.json, beside this script by default) is
LOCAL ONLY: never commit it, never upload it, do back it up.

Usage:
  python3 record_anonymizer.py RECORDS.pdf [--expect N]   # -> RECORDS_ANON.pdf
  python3 record_anonymizer.py --restore OUTPUT.xlsx      # -> OUTPUT_RESTORED.xlsx
  python3 record_anonymizer.py --selftest                 # pure-function tests
  python3 record_anonymizer.py                            # GUI (if tkinter present)

Dependencies: PyMuPDF (pip install pymupdf); openpyxl only for xlsx restore;
tkinterdnd2 optional for drag-and-drop.
"""

from __future__ import annotations

import argparse
import csv
import datetime
import json
import os
import re
import sys

__version__ = "2.1.0"

# --------------------------------------------------------------------------
# Section 0 — constants
# --------------------------------------------------------------------------

COLUMN_SPLIT_X = 304.0        # two-column newspaper flow splits here
BAND_TOL = 3.0                # words within this y-distance form one text row
DETACHED_MAX_BANDS = 6        # how many rows forward a detached/wrapped
                               # value search may walk before giving up

RE_OSIS = re.compile(r"^\d{9}$")
RE_PHONE = re.compile(r"^\d{7,}$")
RE_DATE = re.compile(r"^\d{2}/\d{2}/\d{4}$")
RE_CREDITS = re.compile(r"^\d+\.\d{2}/\d+\.\d{2}$")
RE_YEAR = re.compile(r"^(19|20)\d{2}$")

# Mark vocabulary (handoff §2). "Strong" marks anchor teacher segmentation.
MARK_CODES = {"CR", "NC", "INC", "ABS", "NS", "NU", "NW", "NX", "NL", "ND",
              "ME", "MP", "MA", "MU", "MB", "MT", "WA", "WG", "P", "F"}
RE_MARK_NUMERIC = re.compile(r"^\d{1,3}\*{0,2}$")
RE_MARK_LETTER = re.compile(r"^[A-F][+\-]?\*{0,2}$")
# Unicode-aware: teacher/parent surnames commonly carry accents (MUÑOZ,
# GARCÍA, PEÑA) — an ASCII-only token class silently dropped those from
# every downstream check (mark-anchor gap test, name-shape scans).
RE_NAME_TOKEN = re.compile(r"^[^\W\d_]+(?:[\-'’.][^\W\d_]+)*$")

# Synthetic alias ranges (§6): structurally disjoint (different leading
# digit) so no truncation or substring extraction of one can ever equal a
# value from the other or collide with a real 9-digit OSIS/10-digit phone.
ALIAS_ID_BASE = 900000000     # real OSIS start with 2
ALIAS_PHONE_BASE = 8000000000

BUILTIN_NONPII_PHRASES = (
    "Student Permanent Record", "Admit Date", "Discharge Date",
    "Graduation Date", "Grade Level", "Cumulative Average",
    "Actual Credits", "Credits Earned",
)


def _lbl(*words):
    """Both a spaced ('Word :') and colon-attached ('Word:') label variant.

    PyMuPDF's word tokenization follows the source PDF's exact whitespace;
    real exports may attach the colon to the last label word instead of
    spacing it, and earlier versions of this tool only matched one shape
    (the missing 'Counselor:' variant let populated counselor values pass
    through completely unharvested).
    """
    spaced = list(words) + [":"]
    attached = list(words[:-1]) + [words[-1] + ":"]
    return [spaced, attached]


HEADER_LABELS = {
    # label key -> list of token-sequence variants (words in x order)
    "name_id":    [["Name", "/", "ID", ":"]],
    "address":    _lbl("Address"),
    "phone":      _lbl("Ph#"),
    "dob":        _lbl("DOB"),
    "parent":     [["Parent:"], ["Parent", ":"]],
    "counselor":  _lbl("Counselor"),
    # non-PII labels, matched as value terminators AND compared post-redaction
    # to catch a redaction that clipped a neighboring field:
    "ofcl":       _lbl("Ofcl"),
    "admit":      _lbl("Admit", "Date"),
    "discharge":  _lbl("Discharge", "Date"),
    "graduation": _lbl("Graduation", "Date"),
    "grade":      _lbl("Grade", "Level"),
    "status":     _lbl("Status"),
    "cumulative": _lbl("Cumulative"),
    "cum_avg":    _lbl("Cumulative", "Average"),
}
PII_LABELS = ("name_id", "address", "phone", "dob", "parent", "counselor")
NONPII_COMPARE_KEYS = ("ofcl", "admit", "discharge", "graduation", "grade",
                       "status", "cum_avg")


# --------------------------------------------------------------------------
# Section 1 — alias allocation / mapping store
# --------------------------------------------------------------------------

def student_aliases(idx):
    """Deterministic alias set for student index idx (1-based).

    Alias spec (§6): uppercase 'LAST, FIRST'; first-name alias has a unique
    5-char prefix (F{idx:03d}X...); 9-digit IDs from 900000001; phones from
    8000000001 — a different leading digit than the ID base so no
    truncation or substring of a phone alias can ever equal an ID alias
    (see selftest for the concrete collision this replaced); DOB in
    1900-1909 (never a real student DOB) and unique per idx; address a
    single synthetic token; parent in a disjoint PL/PF namespace.
    """
    if not 1 <= idx <= 999:
        raise ValueError("student idx out of range 1..999: %r" % idx)
    i = idx - 1
    dob = "%02d/%02d/%d" % (1 + (i // 28) % 12, 1 + i % 28, 1900 + i // 336)
    return {
        "last": "L%03dXX" % idx,
        "first": "F%03dXX" % idx,
        "name": "L%03dXX, F%03dXX" % (idx, idx),
        "id": str(ALIAS_ID_BASE + idx),
        "phone": str(ALIAS_PHONE_BASE + idx),
        "dob": dob,
        "address": "ADDR%03dX" % idx,
        "parent": "PL%03dXX, PF%03dXX" % (idx, idx),
    }


def teacher_alias(idx):
    return "TCH%03d" % idx      # single word: fits downstream teacher regex


def counselor_alias(idx):
    return "C%02dCNSL" % idx    # CNSL suffix (§6): disjoint from F/L namespace


class Mapping:
    """anonymizer_mapping.json — deterministic, stable across runs and files.

    Students are keyed by real OSIS ID; teachers/counselors by their exact
    harvested string. LOCAL ONLY: contains real PII by design.
    """

    def __init__(self, path):
        self.path = path
        self.data = {"version": 2, "created": None, "updated": None,
                     "students": {}, "teachers": {}, "counselors": {}}
        if os.path.exists(path):
            with open(path, "r", encoding="utf-8") as fh:
                self.data = json.load(fh)
            if self.data.get("version") != 2:
                raise SystemExit(
                    "Mapping %s has version %r; this tool writes version 2. "
                    "Move the old file aside first." % (path, self.data.get("version")))

    # -- allocation ---------------------------------------------------------

    def _next_student_idx(self):
        used = {s["idx"] for s in self.data["students"].values()}
        idx = 1
        while idx in used:
            idx += 1
        return idx

    def student(self, osis, name, extra):
        """Return (record, aliases) for a real OSIS, allocating if new.

        extra: dict of real field values seen on this page (dob, phone,
        address, parent, counselor) — stored/refreshed for the restore
        direction, but the alias assignment never changes once made.

        Known limitation (mapping evolution): if this OSIS was already
        mapped with different real field values (e.g. a later export shows
        a new address), those values are overwritten here. Restoring an
        OLDER output against the UPDATED mapping will then substitute the
        newer real value. There is no per-value history — back up the
        mapping file before re-running against a materially different
        export if old outputs may still need restoring.
        """
        rec = self.data["students"].get(osis)
        if rec is None:
            idx = self._next_student_idx()
            rec = {"idx": idx, "name": name, "real": {}}
            self.data["students"][osis] = rec
        rec["name"] = rec.get("name") or name
        for k, v in extra.items():
            if v:
                rec["real"][k] = v
        return rec, student_aliases(rec["idx"])

    def teacher(self, name):
        """Alias for a teacher string, preserving a trailing initial.

        'O'BRIEN M' -> 'TCH003 M' rather than 'TCH003': the downstream
        parser's right-to-left fallbacks capture a trailing single letter as
        the mark on modifier-mark rows ('B+', '92**'), so both the token
        signature AND the letter must survive for the anonymized file to
        parse byte-identically. The surname is always fully aliased; a bare
        initial in a course row identifies nobody.
        """
        t = self.data["teachers"]
        if name not in t:
            toks = name.split()
            alias = teacher_alias(len(t) + 1)
            if len(toks) > 1 and len(toks[-1]) == 1:
                alias += " " + toks[-1]
            t[name] = alias
        return t[name]

    def counselor(self, name):
        c = self.data["counselors"]
        if name not in c:
            c[name] = counselor_alias(len(c) + 1)
        return c[name]

    # -- persistence --------------------------------------------------------

    def save(self):
        now = datetime.datetime.now().isoformat(timespec="seconds")
        self.data["created"] = self.data.get("created") or now
        self.data["updated"] = now
        # NOT "*mapping*.json" pattern below: a crash between writing this
        # and os.replace would strand real PII in a file .gitignore doesn't
        # cover. Named so the existing 'anonymizer_mapping.json' glob AND a
        # dedicated '*.tmp' rule both catch it (belt and suspenders).
        tmp = self.path + ".tmp"
        with open(tmp, "w", encoding="utf-8") as fh:
            json.dump(self.data, fh, indent=2, sort_keys=True)
        os.replace(tmp, self.path)

    # -- restore direction --------------------------------------------------

    def reverse_pairs(self):
        """(alias, real) pairs, longest alias first — dictionary restore."""
        pairs = []
        for osis, rec in self.data["students"].items():
            al = student_aliases(rec["idx"])
            pairs.append((al["name"], rec["name"]))
            if "," in rec["name"]:
                last, first = [p.strip() for p in rec["name"].split(",", 1)]
                pairs.append((al["last"], last))
                pairs.append((al["first"], first))
            pairs.append((al["id"], osis))
            real = rec.get("real", {})
            for key, akey in (("dob", "dob"), ("phone", "phone"),
                              ("address", "address")):
                if real.get(key):
                    pairs.append((al[akey], real[key]))
            if real.get("parent"):
                pval = real["parent"]
                pairs.append((al["parent"], pval))
                pl, pf = al["parent"].split(", ")
                # Split like the student name so a BARE half-token restores
                # to just that half, not the whole comma-bearing string
                # (which both introduces an unintended comma into whatever
                # field held the bare token and duplicates the surname).
                if "," in pval:
                    plast, pfirst = [p.strip() for p in pval.split(",", 1)]
                    pairs.append((pl, plast))
                    pairs.append((pf, pfirst))
                else:
                    ptoks = pval.split()
                    if len(ptoks) >= 2:
                        pairs.append((pl, ptoks[-1]))
                        pairs.append((pf, " ".join(ptoks[:-1])))
                    else:
                        pairs.append((pl, pval))
                        pairs.append((pf, pval))
        for name, alias in self.data["teachers"].items():
            pairs.append((alias, name))
            if " " in alias:     # 'TCH001 M' — also map the bare base token
                pairs.append((alias.split()[0], name.rsplit(None, 1)[0]))
        for name, alias in self.data["counselors"].items():
            pairs.append((alias, name))
        pairs.sort(key=lambda p: -len(p[0]))
        return pairs

    def all_real_values(self):
        """Every real value that must be absent from anonymized output."""
        vals = []
        for osis, rec in self.data["students"].items():
            vals.append(rec["name"])
            vals.append(osis)
            vals.extend(v for v in rec.get("real", {}).values() if v)
        vals.extend(self.data["teachers"])
        vals.extend(self.data["counselors"])
        return [v for v in vals if v and len(v) >= 2]

    def alias_id_set(self):
        return {student_aliases(r["idx"])["id"] for r in self.data["students"].values()}

    def alias_phone_set(self):
        return {student_aliases(r["idx"])["phone"] for r in self.data["students"].values()}

    def alias_dob_set(self):
        return {student_aliases(r["idx"])["dob"] for r in self.data["students"].values()}


# --------------------------------------------------------------------------
# Section 2 — word geometry
# --------------------------------------------------------------------------

class W:
    """One extracted word with its box."""
    __slots__ = ("x0", "y0", "x1", "y1", "text")

    def __init__(self, x0, y0, x1, y1, text):
        self.x0, self.y0, self.x1, self.y1, self.text = x0, y0, x1, y1, text

    @property
    def yc(self):
        return (self.y0 + self.y1) / 2.0

    def __repr__(self):
        return "W(%r @%.0f,%.0f)" % (self.text, self.x0, self.y0)


def page_words(page):
    return [W(w[0], w[1], w[2], w[3], w[4]) for w in page.get_text("words")]


def band_cluster(words, tol=BAND_TOL):
    """Group words into horizontal text rows by y-center proximity."""
    bands = []
    for w in sorted(words, key=lambda w: (w.yc, w.x0)):
        if bands and abs(w.yc - bands[-1][0].yc) <= tol:
            bands[-1].append(w)
        else:
            bands.append([w])
    for b in bands:
        b.sort(key=lambda w: w.x0)
    return bands


def union_rect(words):
    return (min(w.x0 for w in words), min(w.y0 for w in words),
            max(w.x1 for w in words), max(w.y1 for w in words))


# --------------------------------------------------------------------------
# Section 3 — header harvest
# --------------------------------------------------------------------------

def _match_label_at(tokens, i):
    """If a label starts at tokens[i], return (key, end_index_exclusive)."""
    for key, variants in HEADER_LABELS.items():
        for seq in variants:
            if [t.text for t in tokens[i:i + len(seq)]] == seq:
                return key, i + len(seq)
    return None, i


def _find_labels(tokens):
    """All label occurrences in a token row: list of (key, start, end)."""
    out, i = [], 0
    while i < len(tokens):
        key, end = _match_label_at(tokens, i)
        if key:
            out.append((key, i, end))
            i = end
        else:
            i += 1
    return out


def _is_course_anchor(band):
    return any(w.text == "Course" and w.x0 < 20 for w in band)


def _forward_value(bands, band_labels, band_idx):
    """Walk forward from a band to collect a detached or wrapped value.

    Unlike a fixed point-distance window, this walks row by row until it
    hits a real structural boundary: another label (any key), the course
    table's 'Course' anchor, the header column running out (no more
    left-column words), or a generous row cap — never a guessed pixel
    tolerance that can silently undershoot a value sitting one row further
    down, or a value that legitimately wraps onto a second physical line.
    """
    collected = []
    for bi in range(band_idx + 1, min(len(bands), band_idx + 1 + DETACHED_MAX_BANDS)):
        band = bands[bi]
        if band_labels[bi] or _is_course_anchor(band):
            break
        left_words = [w for w in band if w.x0 < COLUMN_SPLIT_X]
        if not left_words:
            break
        collected.extend(left_words)
    return collected


def harvest_page(words):
    """Extract PII fields from one page's words.

    Returns dict: field -> {"value": str, "words": [W]}; plus "_labels" with
    the label spans found (for the harvest floor + field-presence checks).
    """
    bands = band_cluster(words)
    band_labels = [_find_labels(b) for b in bands]
    fields = {}
    label_spans = []

    for bi, band in enumerate(bands):
        labels = band_labels[bi]
        blank_pii = []   # [(key, label_last_word)] needing forward resolution
        for n, (key, start, end) in enumerate(labels):
            label_spans.append((key, band[start], band[end - 1]))
            if key not in PII_LABELS:
                continue
            stop = labels[n + 1][1] if n + 1 < len(labels) else len(band)
            same_row = list(band[end:stop])
            if key == "name_id":
                fields["name_id"] = _parse_name_id(same_row)
            elif key == "phone":
                tok = next((w for w in same_row if RE_PHONE.match(w.text)), None)
                if tok:
                    fields["phone"] = {"value": tok.text, "words": [tok]}
            elif key == "dob":
                tok = next((w for w in same_row if RE_DATE.match(w.text)), None)
                if tok:
                    fields["dob"] = {"value": tok.text, "words": [tok]}
            elif key in ("address", "parent", "counselor"):
                if same_row:
                    fields[key] = {"value": " ".join(w.text for w in same_row),
                                   "words": same_row}
                else:
                    blank_pii.append((key, band[end - 1]))
        if blank_pii:
            # A row can hold more than one blank label (e.g. 'Parent:' and
            # an empty 'Counselor :' side by side) — the SAME forward
            # content must not be handed to every one of them. Compute it
            # once for the row and assign it only to whichever blank
            # label's x-position is closest to where the content starts
            # (values are visually indented near their own label).
            forward = _forward_value(bands, band_labels, bi)
            if forward:
                content_x0 = forward[0].x0
                key, _lw = min(blank_pii, key=lambda kl: abs(kl[1].x0 - content_x0))
                fields[key] = {"value": " ".join(w.text for w in forward),
                               "words": forward}
    fields["_labels"] = label_spans
    return fields


def _parse_name_id(value_words):
    """['LASTNAME,', 'FIRSTNAME', '/', '212345678'] -> name + id entries."""
    try:
        slash = next(i for i, w in enumerate(value_words) if w.text == "/")
    except StopIteration:
        return None
    name_words = value_words[:slash]
    id_tok = next((w for w in value_words[slash + 1:] if RE_OSIS.match(w.text)), None)
    if not name_words or id_tok is None:
        return None
    return {"value": " ".join(w.text for w in name_words), "words": name_words,
            "id": id_tok.text, "id_word": id_tok}


def harvest_nonpii(words):
    """Non-PII header fields, for the pre/post-redaction comparison check."""
    bands = band_cluster(words)
    band_labels = [_find_labels(b) for b in bands]
    out = {}
    for bi, band in enumerate(bands):
        labels = band_labels[bi]
        for n, (key, start, end) in enumerate(labels):
            if key not in NONPII_COMPARE_KEYS:
                continue
            stop = labels[n + 1][1] if n + 1 < len(labels) else len(band)
            out.setdefault(key, []).append(
                " ".join(w.text for w in band[end:stop]))
    return out


def field_stats(page_infos):
    """Per-field populated/blank counts across the document, for the report.

    Address/Counselor are documented as legitimately blank on many pages
    (handoff §2), so an empty harvest is not itself a failure — but it must
    never be silent. This makes every field's coverage visible in every
    run's report so a suspicious pattern (e.g. phone blank on all pages)
    is something the user can see and investigate, honoring the 'harvest
    floor' philosophy for fields beyond Name/ID without hard-failing on
    values the spec says are normally absent.
    """
    lines = []
    for key in ("dob", "phone", "address", "parent", "counselor"):
        labeled = sum(1 for i in page_infos if key in i["labels_seen"])
        populated = sum(1 for i in page_infos if i.get("fields_seen", {}).get(key))
        lines.append("field '%s': %d/%d labeled pages populated"
                     % (key, populated, labeled))
    return lines


# --------------------------------------------------------------------------
# Section 4 — course-row harvest (teacher layer)
# --------------------------------------------------------------------------

def _mark_class(tok):
    """0 = not a mark; 1 = strong; 2 = bare letter grade; 3 = bare S/E."""
    if RE_MARK_NUMERIC.match(tok):
        return 1
    core = tok.rstrip("*")
    starred = tok != core
    if core in MARK_CODES:
        return 1
    if RE_MARK_LETTER.match(tok):
        if starred or len(core) == 2:      # 'B*', 'B+' — modifier disambiguates
            return 1
        return 2                            # bare 'B'
    if core in ("S", "E"):
        return 1 if starred else 3          # stray artifacts (handoff §2)
    return 0


def _is_course_row(tokens):
    """Course rows start 'YYYY / T' and end (somewhere) with d.dd/d.dd."""
    return (len(tokens) >= 6 and RE_YEAR.match(tokens[0].text)
            and tokens[1].text == "/" and tokens[2].text.isdigit())


def _row_starts_new_record(tokens):
    """True if tokens open a NEW course or exam row ('YYYY / T' / 'YYYY Term')."""
    if len(tokens) < 2 or not RE_YEAR.match(tokens[0].text):
        return False
    return tokens[1].text in ("/", "Term")


def segment_course_row(tokens):
    """Split a course row's tokens into fields; find the teacher span.

    Returns dict with year/term/school/code, name tokens, mark token,
    teacher tokens, credits token, and 'permissive' flag — or None if the
    row is not a course row. Teacher identification is mark-anchored: among
    candidate mark tokens in the 4 slots before the credits anchor whose gap
    to the credits consists only of name-shaped tokens (max 3), the best
    mark class wins (strong > bare letter > bare S/E); within a class the
    leftmost valid candidate wins, which redacts the most (privacy bias).
    """
    if not _is_course_row(tokens):
        return None
    cred_i = next((i for i in range(len(tokens) - 1, 3, -1)
                   if RE_CREDITS.match(tokens[i].text)), None)
    if cred_i is None or cred_i < 5:
        return None
    lo = max(4, cred_i - 4)
    candidates = []          # (mark_class, index)
    for i in range(lo, cred_i):
        mc = _mark_class(tokens[i].text)
        if not mc:
            continue
        gap = tokens[i + 1:cred_i]
        if len(gap) <= 3 and all(RE_NAME_TOKEN.match(t.text) for t in gap):
            candidates.append((mc, i))
    base = {"year": tokens[0].text, "term": tokens[2].text,
            "school": tokens[3].text, "code": tokens[4].text,
            "credits": tokens[cred_i], "permissive": False}
    if candidates:
        best_class = min(c for c, _ in candidates)
        mark_i = min(i for c, i in candidates if c == best_class)
        base.update(name=tokens[5:mark_i], mark=tokens[mark_i],
                    teacher=tokens[mark_i + 1:cred_i])
        return base
    # No mark anchor found: ambiguous row. Privacy bias — treat every
    # name-shaped token of len>=2 directly before the credits (up to 3) as
    # teacher, redact it, and log the row (handoff §5.4).
    t = []
    for tok in reversed(tokens[max(5, cred_i - 3):cred_i]):
        if RE_NAME_TOKEN.match(tok.text) and len(tok.text) >= 2:
            t.insert(0, tok)
        else:
            break
    base.update(name=tokens[5:cred_i - len(t)], mark=None, teacher=t,
                permissive=True)
    return base


def logical_rows(words):
    """Token rows for course parsing: y-bands split at the column boundary.

    A band is only split into independent left/right rows when BOTH sides
    look like independent content. If the left side looks like the START
    of a course row (a long course name/teacher whose trailing credits
    token happens to land at x>=304) but is missing its credits anchor,
    and the right side does not itself start a new course/exam record,
    the split was purely an artifact of one row's tail crossing the
    column boundary — merge it back into a single row instead of silently
    dropping the teacher redaction on the left fragment and discarding an
    orphaned '1.00/1.00' on the right.
    """
    rows = []
    for band in band_cluster(words):
        left = [w for w in band if w.x0 < COLUMN_SPLIT_X]
        right = [w for w in band if w.x0 >= COLUMN_SPLIT_X]
        if (left and right and _is_course_row(left)
                and not any(RE_CREDITS.match(w.text) for w in left)
                and not _row_starts_new_record(right)):
            rows.append(left + right)
            continue
        for part in (left, right):
            if part:
                rows.append(part)
    return rows


def harvest_courses(words):
    """All course-row segmentations on a page (for teachers + verification)."""
    out = []
    for row in logical_rows(words):
        seg = segment_course_row(row)
        if seg:
            out.append(seg)
    return out


def harvest_exams(words):
    """Exam rows 'YYYY Term N CODE NAME... SCORE' — no PII, used to verify."""
    out = []
    for row in logical_rows(words):
        t = [w.text for w in row]
        if (len(t) >= 5 and RE_YEAR.match(t[0]) and t[1] == "Term"
                and t[2].isdigit() and t[-1].isdigit()):
            out.append(tuple(t))
    return out


# --------------------------------------------------------------------------
# Section 5 — redaction planning and application
# --------------------------------------------------------------------------

class Redaction:
    __slots__ = ("rect", "alias", "kind")

    def __init__(self, rect, alias, kind):
        self.rect, self.alias, self.kind = rect, alias, kind


def _rects_overlap(a, b):
    return not (a[2] <= b[0] or b[2] <= a[0] or a[3] <= b[1] or b[3] <= a[1])


def _rect_contains(outer, inner, pad=0.75):
    return (outer[0] - pad <= inner[0] and outer[1] - pad <= inner[1]
            and outer[2] + pad >= inner[2] and outer[3] + pad >= inner[3])


def _inset(rect):
    """Shrink a redaction rect so it cannot clip neighboring lines.

    PyMuPDF's apply_redactions removes every character whose box merely
    TOUCHES the rect; word boxes of adjacent text rows can overlap by
    sub-point margins, which silently deleted neighboring non-PII text
    (e.g. an 'Ofcl :' field) in testing. The redacted value's own chars
    span the full rect height, so an inset rect still removes all of them.
    """
    x0, y0, x1, y1 = rect
    dy = min(1.2, (y1 - y0) * 0.2)
    return (x0 + 0.2, y0 + dy, x1 - 0.2, y1 - dy)


def plan_page(page_no, words, mapping, log):
    """Harvest one page and plan its redactions.

    Returns (redactions, page_info). page_info feeds the harvest floor and
    the verification report.
    """
    fields = harvest_page(words)
    labels_seen = {k for k, _, _ in fields["_labels"]}
    info = {"page": page_no,
            "has_name_label": "name_id" in labels_seen,
            "labels_seen": labels_seen, "fields_seen": {},
            "osis": None, "aliases": {}, "permissive_rows": []}
    red = []

    ni = fields.get("name_id")
    if ni:
        extra = {k: fields[k]["value"] for k in
                 ("dob", "phone", "address", "parent", "counselor") if k in fields}
        rec, al = mapping.student(ni["id"], ni["value"], extra)
        info["osis"] = ni["id"]
        # One combined redaction for "NAME / ID": short real names would not
        # leave enough width to draw the alias before the '/' otherwise.
        name_id_alias = "%s / %s" % (al["name"], al["id"])
        info["aliases"]["name_id"] = name_id_alias
        red.append(Redaction(union_rect(ni["words"] + [ni["id_word"]]),
                             name_id_alias, "name_id"))
        for key in ("dob", "phone", "address", "parent"):
            if key in fields:
                info["fields_seen"][key] = True
                info["aliases"][key] = al[key]
                red.append(Redaction(union_rect(fields[key]["words"]),
                                     al[key], key))
        if "counselor" in fields:
            info["fields_seen"]["counselor"] = True
            alias = mapping.counselor(fields["counselor"]["value"])
            info["aliases"]["counselor"] = alias
            red.append(Redaction(union_rect(fields["counselor"]["words"]),
                                 alias, "counselor"))
    elif info["has_name_label"]:
        log.append("page %d: Name / ID label present but value not parseable"
                   % page_no)

    for seg in harvest_courses(words):
        if seg["teacher"]:
            tname = " ".join(w.text for w in seg["teacher"])
            alias = mapping.teacher(tname)
            red.append(Redaction(union_rect(seg["teacher"]), alias, "teacher"))
            if seg["permissive"]:
                info["permissive_rows"].append(
                    "page %d: no mark anchor; permissively redacted a "
                    "%d-token span in a course row (name masked: %s...)"
                    % (page_no, len(seg["teacher"]), tname[:1] + "*" * (len(tname) - 1)))
    return red, info


def _room_bound(rect, words, exclude_rects, page_width):
    """How far right of rect a replacement may extend before hitting the
    next surviving word (or the page edge).

    This bound is used for BOTH the redaction rect AND the alias draw so
    they always agree: an earlier version redacted only the original
    word's own (narrow) box, then drew a WIDER alias into the gap that
    followed it. That gap still held the original inter-word space
    glyph(s) (PyMuPDF lays out real space characters with their own
    bounding boxes), which PyMuPDF's own word extraction ignored but a
    position-based extractor (pdfplumber, used by the downstream Excel
    pipeline) treated as a genuine word-break sitting INSIDE the visually
    intact alias — splitting e.g. 'TCH008' into 'TCH00' + '8'. Redacting
    through to the same boundary the draw uses removes that leftover
    glyph regardless of how much wider the alias turns out to be.
    """
    x0, y0, x1, y1 = rect
    yc = (y0 + y1) / 2
    obstacles = [w.x0 for w in words
                 if w.x0 > x1 - 0.5 and abs(w.yc - yc) < BAND_TOL
                 and not any(_rects_overlap((w.x0, w.y0, w.x1, w.y1), rr)
                             for rr in exclude_rects)]
    if obstacles:
        return min(obstacles) - 1.0
    return min(x1 + 80.0, page_width - 6.0)


def apply_redactions(doc, plans, log):
    """Redact (text removal) then draw aliases with a font-size fit loop.

    Aliases are drawn with insert_text after apply_redactions so that the
    replacement text is never silently clipped: if an alias cannot be drawn
    even at the minimum font size an ERROR is logged AND the specific
    field/page is recorded as an unfilled alias — the alias-presence check
    in verify() then fails the run for every PII kind, not just name/ID.
    """
    import fitz
    for page, reds, words in plans:
        if not reds:
            continue
        orig_rects = [r.rect for r in reds]
        room_bounds = [_room_bound(r.rect, words, orig_rects, page.rect.width)
                       for r in reds]
        for r, x1 in zip(reds, room_bounds):
            x0, y0, _, y1 = r.rect
            page.add_redact_annot(fitz.Rect(*_inset((x0, y0, x1, y1))))
        try:
            page.apply_redactions(graphics=0)
        except TypeError:                      # older PyMuPDF signature
            page.apply_redactions()
        for r, x1 in zip(reds, room_bounds):
            _draw_alias(page, r, x1, log)


def _draw_alias(page, red, max_x1, log):
    """Draw the alias at the redacted value's baseline, shrinking to fit.

    If nothing fits even at 4pt, log an ERROR — the alias-presence check in
    verify() then fails the run rather than shipping a blank field.
    """
    import fitz
    x0, y0, x1, y1 = red.rect
    room = max_x1 - x0
    for fs in (8.0, 7.0, 6.0, 5.0, 4.0):
        if fitz.get_text_length(red.alias, fontname="helv", fontsize=fs) <= room:
            page.insert_text((x0, y1 - 1.2), red.alias, fontsize=fs,
                             fontname="helv")
            return
    log.append("ERROR: alias for kind=%r does not fit at %s (drawn nothing "
              "at this position)" % (red.kind, red.rect))


# --------------------------------------------------------------------------
# Section 6 — verification
# --------------------------------------------------------------------------

def _tolerant_pattern(value):
    """Whitespace-tolerant regex for a harvested value string."""
    return re.compile(r"\s+".join(re.escape(t) for t in value.split()))


def _mask(value, keep=1):
    """Report-safe rendering of a real value: never print it verbatim.

    Verification-failure reports are exactly what a user is likely to paste
    into a support request when a run FAILs — they must not themselves leak
    the PII the tool exists to protect.
    """
    if not value:
        return value
    return value[:keep] + "*" * max(0, len(value) - keep)


_SURNAME = r"[^\W\d_][^\W\d_\-'’]*(?:[\-'’][^\W\d_]+)*"   # min 2 chars
_GIVEN = r"[^\W\d_][^\W\d_\-'’]*\.?"                          # min 1 char
RE_SCAN_NAME = re.compile(
    r"\b%s(?: %s)*,\s+%s\b" % (_SURNAME, _SURNAME, _GIVEN))
# 'First Last' mixed-case order (documented real Parent shape): 2-4
# consecutive Title-Case words, no comma required.
RE_SCAN_NAME_TITLE = re.compile(r"\b[A-Z][a-z]+(?: [A-Z][a-z]+){1,3}\b")
RE_SCAN_DATE = re.compile(r"\b\d{2}/\d{2}/((?:19|20)\d{2})\b")
RE_SCAN_ID9 = re.compile(r"\b\d{9}\b")
RE_SCAN_ID10 = re.compile(r"\b\d{10}\b")
RE_SCAN_STREET = re.compile(
    r"\b\d{1,5} (?:[A-Z0-9'\-]+ ){1,4}"
    r"(?:ST|STREET|AVE|AVENUE|RD|ROAD|BLVD|BOULEVARD|PL|PLACE|CT|COURT|"
    r"LN|LANE|DR|DRIVE|PKWY|PARKWAY|TER|TERRACE|WAY|LOOP)\b")
RE_SCAN_ZIP = re.compile(r"\b1[01]\d{3}(?:-\d{4})?\b")
RE_ALIAS_NAMEISH = re.compile(r"^P?L\d{3}XX, P?F\d{3}XX$")
# Any alias-SHAPED token, across every namespace — used as the residual
# check on restore output (an alias that survives restore, whatever the
# reason, is a data-quality bug the user must see, not a silent gap).
RE_ANY_ALIAS_SHAPE = re.compile(
    r"\bP?L\d{3}XX\b|\bP?F\d{3}XX\b|\bTCH\d{3}(?: [A-Za-z])?\b|"
    r"\bC\d{2}CNSL\b|\bADDR\d{3}X\b|\b[89]\d{8,9}\b")


def _scrub_builtin(text):
    for p in BUILTIN_NONPII_PHRASES:
        text = text.replace(p, " " * len(p))
    return text


def verify(anon_path, mapping, page_infos, allow_patterns, log):
    """Post-anonymization verification. Returns (ok, failures, warnings)."""
    import fitz
    failures, warnings = [], []

    doc = fitz.open(anon_path)
    texts = [p.get_text() for p in doc]
    metadata = dict(doc.metadata or {})
    toc_titles = [t[1] for t in doc.get_toc()]
    doc.close()
    full = "\n".join(texts)

    # (1) Residual scan for every known real value (whole file, all pages),
    #     including PDF metadata/bookmarks (Title/Author/Subject/Keywords,
    #     outline entries) — a redaction pass that only touches page text
    #     would otherwise ship real PII in the Info dictionary untouched.
    #     A case-INSENSITIVE hit is also a FAIL, not a warning: a value the
    #     tool has positively identified as real PII is not an ambiguity
    #     just because its capitalization differs somewhere in the file.
    for value in mapping.all_real_values():
        pat = _tolerant_pattern(value)
        ipat = re.compile(pat.pattern, re.IGNORECASE)
        for pno, text in enumerate(texts, 1):
            if pat.search(text):
                failures.append("page %d: real value still present (masked: %s)"
                                % (pno, _mask(value)))
            elif ipat.search(text):
                failures.append("page %d: case-variant of a real value present "
                                "(masked: %s)" % (pno, _mask(value)))
        for field, mval in metadata.items():
            if mval and (pat.search(str(mval)) or ipat.search(str(mval))):
                failures.append("metadata field %r still contains a real "
                                "value (masked: %s)" % (field, _mask(value)))
        for title in toc_titles:
            if pat.search(title) or ipat.search(title):
                failures.append("bookmark/outline title still contains a "
                                "real value (masked: %s)" % _mask(value))

    # (2) Alias presence: every PII field harvested on a page must show its
    #     alias in the output — covers name/ID AND dob/phone/address/
    #     parent/counselor, so a font-fit failure on any of them fails the
    #     run instead of silently shipping a blank field under a PASS.
    for info in page_infos:
        for kind, alias in info["aliases"].items():
            text = texts[info["page"] - 1]
            if not _tolerant_pattern(alias).search(text):
                failures.append("page %d: alias for %s missing from output "
                                "(redaction may not have drawn)"
                                % (info["page"], kind))

    # (3) Pattern-class scan (§5.3): shapes, not just known values.
    alias_ids = mapping.alias_id_set()
    alias_phones = mapping.alias_phone_set()
    alias_dobs = mapping.alias_dob_set()
    cur_year = datetime.date.today().year
    for pno, text in enumerate(texts, 1):
        scrubbed = _scrub_builtin(text)
        for m in RE_SCAN_NAME.finditer(scrubbed):
            s = m.group(0)
            if RE_ALIAS_NAMEISH.match(s) or s.strip() in allow_patterns:
                continue
            failures.append("page %d: 'LAST, FIRST'-shaped string outside "
                            "the alias namespace (masked: %s)"
                            % (pno, _mask(s)))
        for m in RE_SCAN_NAME_TITLE.finditer(scrubbed):
            s = m.group(0)
            if s.strip() in allow_patterns:
                continue
            failures.append("page %d: Title-Case 'First Last'-shaped string "
                            "(masked: %s)" % (pno, _mask(s)))
        for m in RE_SCAN_DATE.finditer(text):
            year = int(m.group(1))
            if m.group(0) in alias_dobs:
                continue
            line = text[text.rfind("\n", 0, m.start()) + 1:
                        max(text.find("\n", m.end()), m.end())]
            if re.search(r"(Admit|Discharge|Graduation)\s*Date", line):
                continue
            if 1990 <= year <= cur_year - 8:
                failures.append("page %d: DOB-range date %s outside alias set"
                                % (pno, m.group(0)))
        for m in RE_SCAN_ID9.finditer(text):
            if m.group(0) not in alias_ids and not _inside_longer_digits(text, m):
                failures.append("page %d: 9-digit number %s outside synthetic "
                                "range" % (pno, m.group(0)))
        for m in RE_SCAN_ID10.finditer(text):
            if m.group(0) not in alias_phones:
                failures.append("page %d: 10-digit number %s outside synthetic "
                                "range" % (pno, m.group(0)))
        for m in RE_SCAN_STREET.finditer(scrubbed):
            if m.group(0).strip() in allow_patterns:
                continue
            failures.append("page %d: street-address-shaped string (masked: "
                            "%s)" % (pno, _mask(m.group(0))))
        for m in RE_SCAN_ZIP.finditer(text):
            warnings.append("page %d: NYC-zip-shaped number %s (single-token "
                            "coincidence?)" % (pno, m.group(0)))

    ok = not failures
    return ok, failures, warnings


def _inside_longer_digits(text, m):
    before = text[m.start() - 1:m.start()]
    after = text[m.end():m.end() + 1]
    return before.isdigit() or after.isdigit()


def check_floor(page_infos, expect_students, log):
    """Harvest-count floor (§5.2): hard FAIL on any shortfall."""
    failures = []
    labeled = [i for i in page_infos if i["has_name_label"]]
    harvested = [i for i in labeled if i["osis"]]
    if len(harvested) != len(labeled):
        failures.append("harvest floor: %d pages carry a Name / ID label but "
                        "only %d were harvested"
                        % (len(labeled), len(harvested)))
    if not labeled:
        failures.append("harvest floor: no Name / ID labels found at all — "
                        "wrong document format?")
    unique = {i["osis"] for i in harvested if i["osis"]}
    if expect_students is not None and len(unique) != expect_students:
        failures.append("harvest floor: expected %d unique students, "
                        "harvested %d" % (expect_students, len(unique)))
    log.append("harvest: %d labeled pages, %d harvested, %d unique students"
               % (len(labeled), len(harvested), len(unique)))
    return failures, len(unique)


def _course_row_lists(words):
    """Course-shaped token rows in page order (loose detection, no segment)."""
    rows = []
    for row in logical_rows(words):
        if _is_course_row(row) and any(RE_CREDITS.match(w.text) for w in row):
            rows.append(row)
    return rows


def compare_course_data(orig_doc, anon_doc, mapping, log):
    """Invariant (c): every course row's tokens are identical apart from the
    teacher tokens, which must equal that teacher's alias; exam rows must be
    byte-identical. Permissive rows are excluded (they are logged instead).
    """
    failures = []
    skipped = 0
    if len(anon_doc) != len(orig_doc):
        failures.append("page count changed: %d -> %d"
                        % (len(orig_doc), len(anon_doc)))
    for pno in range(min(len(anon_doc), len(orig_doc))):
        ow, aw = page_words(orig_doc[pno]), page_words(anon_doc[pno])
        orows, arows = _course_row_lists(ow), _course_row_lists(aw)
        if len(orows) != len(arows):
            failures.append("page %d: %d course rows became %d after "
                            "anonymization" % (pno + 1, len(orows), len(arows)))
            continue
        for orow, arow in zip(orows, arows):
            seg = segment_course_row(orow)
            if seg is None:
                continue
            if seg["permissive"]:
                skipped += 1
                continue
            expected = [w.text for w in orow]
            if seg["teacher"]:
                tname = " ".join(w.text for w in seg["teacher"])
                alias = mapping.data["teachers"].get(tname)
                if alias is None:
                    failures.append("page %d: teacher in a course row was "
                                    "never aliased" % (pno + 1))
                    continue
                ti = orow.index(seg["teacher"][0])
                tj = orow.index(seg["teacher"][-1])
                expected = (expected[:ti] + alias.split()
                            + [w.text for w in orow[tj + 1:]])
            if [w.text for w in arow] != expected:
                failures.append("page %d: course-row tokens changed beyond "
                                "the teacher alias" % (pno + 1))
        if harvest_exams(ow) != harvest_exams(aw):
            failures.append("page %d: exam rows differ after anonymization"
                            % (pno + 1))
    if skipped:
        log.append("course-data compare: %d permissive row(s) excluded"
                   % skipped)
    return failures


def compare_header_fields(orig_doc, anon_doc, log):
    """Invariant (c) for the header: non-PII fields must survive redaction
    byte-identically. This is exactly the failure class _inset() exists
    for (a redaction once silently deleted a neighboring 'Ofcl :' value in
    testing) — generalized into a real per-run guarantee instead of only
    being checked by the synthetic test suite.
    """
    failures = []
    for pno in range(min(len(orig_doc), len(anon_doc))):
        oh = harvest_nonpii(page_words(orig_doc[pno]))
        ah = harvest_nonpii(page_words(anon_doc[pno]))
        for key in NONPII_COMPARE_KEYS:
            if oh.get(key) != ah.get(key):
                failures.append("page %d: non-PII field %r changed after "
                                "anonymization" % (pno + 1, key))
    return failures


# --------------------------------------------------------------------------
# Section 7 — anonymize driver
# --------------------------------------------------------------------------

def default_mapping_path():
    return os.path.join(os.path.dirname(os.path.abspath(__file__)),
                        "anonymizer_mapping.json")


def anonymize_pdf(pdf_path, mapping_path=None, expect_students=None,
                  allow_patterns=(), report_path=None):
    """Full pipeline. Returns (ok, out_path, report_lines).

    Output is written under a '.partial' name first and only renamed to
    the clean '_ANON.pdf' name after every check has PASSed; any exception
    (including Ctrl-C) during redaction/verification leaves either nothing
    or an '_ANON_FAILED.pdf'-suffixed file on disk — never a file under the
    clean name that has not actually been verified.
    """
    import fitz
    log = []
    mapping = Mapping(mapping_path or default_mapping_path())
    doc = fitz.open(pdf_path)

    plans, page_infos = [], []
    for pno, page in enumerate(doc, 1):
        words = page_words(page)
        reds, info = plan_page(pno, words, mapping, log)
        plans.append((page, reds, words))
        page_infos.append(info)
        log.extend(info["permissive_rows"])

    failures, n_students = check_floor(page_infos, expect_students, log)
    log.extend(field_stats(page_infos))

    base, _ = os.path.splitext(pdf_path)
    final_pass_path = base + "_ANON.pdf"
    final_fail_path = base + "_ANON_FAILED.pdf"
    work_path = base + "_ANON.pdf.partial"

    if failures:
        # Hard FAIL before writing anything uploadable.
        report = _report(pdf_path, None, False, failures, [], log, n_students)
        _write_report(report_path or base + "_ANON_report.txt", report)
        doc.close()
        return False, None, report

    ok = False
    vwarn = []
    out_path = None
    try:
        apply_redactions(doc, plans, log)
        doc.save(work_path, garbage=3, deflate=True)
        doc.close()

        ok, vfail, vwarn = verify(work_path, mapping, page_infos,
                                  list(allow_patterns), log)
        orig = fitz.open(pdf_path)
        anon = fitz.open(work_path)
        cfail = compare_course_data(orig, anon, mapping, log)
        hfail = compare_header_fields(orig, anon, log)
        orig.close()
        anon.close()
        if cfail or hfail:
            ok = False
        failures = vfail + cfail + hfail
    except (KeyboardInterrupt, SystemExit):
        ok = False
        failures = failures + ["interrupted before verification completed"]
        raise
    except Exception as exc:
        ok = False
        failures = failures + ["EXCEPTION during anonymization: %s: %s"
                               % (type(exc).__name__, exc)]
    finally:
        try:
            if not doc.is_closed:
                doc.close()
        except Exception:
            pass
        if os.path.exists(work_path):
            if ok:
                os.replace(work_path, final_pass_path)
                out_path = final_pass_path
            else:
                os.replace(work_path, final_fail_path)
                out_path = final_fail_path
        n_red = sum(len(r) for _, r, _ in plans)
        log.append("redactions placed: %d" % n_red)
        report = _report(pdf_path, out_path, ok, failures, vwarn, log, n_students)
        _write_report(report_path or base + "_ANON_report.txt", report)
        if ok:
            mapping.save()

    return ok, out_path, report


def _report(src, out, ok, failures, warnings, log, n_students):
    lines = ["record_anonymizer v%s — %s" % (__version__, "PASS" if ok else "FAIL"),
             "source: %s" % os.path.basename(src),
             "output: %s" % (os.path.basename(out) if out else "(none — failed "
                             "before writing output)"),
             "unique students: %d" % n_students, ""]
    if failures:
        lines.append("FAILURES (%d) — DO NOT UPLOAD THE OUTPUT:" % len(failures))
        lines.extend("  ! " + f for f in failures)
        lines.append("")
    if warnings:
        lines.append("warnings (%d):" % len(warnings))
        lines.extend("  ~ " + w for w in warnings)
        lines.append("")
    lines.append("log:")
    lines.extend("  . " + entry for entry in log)
    return lines


def _write_report(path, lines):
    with open(path, "w", encoding="utf-8") as fh:
        fh.write("\n".join(lines) + "\n")


# --------------------------------------------------------------------------
# Section 8 — restore
# --------------------------------------------------------------------------

def _build_restorer(mapping):
    pairs = mapping.reverse_pairs()
    if not pairs:
        raise SystemExit("mapping is empty — nothing to restore")
    # \b-anchored: an earlier version matched alias substrings anywhere,
    # so a 9-digit ID alias could match inside an unrelated longer number
    # (a computed value, a timestamp) and corrupt it.
    pat = re.compile("|".join(r"\b%s\b" % re.escape(a) for a, _ in pairs))
    table = dict(pairs)
    return pat, table


def restore_text(text, pat, table):
    return pat.sub(lambda m: table[m.group(0)], text)


def _residual_alias_hits(text):
    return sorted(set(RE_ANY_ALIAS_SHAPE.findall(text)))


def restore_file(path, mapping_path=None):
    """Restore real values into a returned txt/csv/xlsx/pdf.

    Returns (out_path, issues). issues is a list of human-readable strings
    covering BOTH failure classes the restore direction can hit silently:
    an alias-shaped token still present afterward (didn't match — stale
    mapping, case mismatch, or format the tool doesn't parse), and — PDF
    only — a real value that could not be DRAWN because it didn't fit in
    the space the alias occupied, which leaves neither the alias nor the
    real value on the page. An empty list means a clean restore; callers
    (CLI/GUI) must surface a non-empty list prominently, not bury it in a
    log stream, since this is exactly the moment a user might miss that
    part of their document silently didn't come back.
    """
    mapping = Mapping(mapping_path or default_mapping_path())
    pat, table = _build_restorer(mapping)
    base, ext = os.path.splitext(path)
    out = base + "_RESTORED" + ext
    ext_l = ext.lower()
    if ext_l in (".txt", ".md", ".json", ".tsv"):
        with open(path, "r", encoding="utf-8", errors="surrogateescape") as fh:
            data = fh.read()
        restored = restore_text(data, pat, table)
        with open(out, "w", encoding="utf-8", errors="surrogateescape") as fh:
            fh.write(restored)
        issues = _residual_issues(_residual_alias_hits(restored))
    elif ext_l == ".csv":
        issues = _restore_csv(path, out, pat, table)
    elif ext_l == ".xlsx":
        issues = _restore_xlsx(path, out, pat, table)
    elif ext_l == ".pdf":
        out, issues = _restore_pdf(path, out, pat, table)
    else:
        raise SystemExit("don't know how to restore %r files" % ext)
    return out, issues


def _residual_alias_hits(text):
    return sorted(set(RE_ANY_ALIAS_SHAPE.findall(text)))


def _residual_issues(hits):
    if not hits:
        return []
    return ["%d alias-shaped token(s) remain after restore (stale mapping, "
           "case mismatch, or an unsupported cell format?): %s"
           % (len(hits), hits[:8])]


def _restore_csv(path, out, pat, table):
    """CSV-aware: substitute per cell, then re-serialize with quoting.

    A plain text substitution would insert an unquoted comma wherever an
    alias restores to a comma-bearing real value ('LAST, FIRST' names,
    'C01CNSL' -> a counselor string), shifting every later column. Reading
    and re-writing through the csv module lets the writer quote any field
    that now needs it.
    """
    with open(path, "r", newline="", encoding="utf-8", errors="surrogateescape") as fh:
        rows = list(csv.reader(fh))
    restored_rows = [[restore_text(cell, pat, table) for cell in row] for row in rows]
    with open(out, "w", newline="", encoding="utf-8", errors="surrogateescape") as fh:
        csv.writer(fh).writerows(restored_rows)
    flat = "\n".join(",".join(r) for r in restored_rows)
    return _residual_issues(_residual_alias_hits(flat))


def _restore_xlsx(path, out, pat, table):
    """Cell-by-cell (§6): never via LLM. Handles non-string cell types —
    spreadsheet writers (pandas/Excel/openpyxl) commonly store all-digit
    alias IDs/phones as numbers and DOB aliases as dates, not strings; a
    string-only restore silently left those untouched.
    """
    import openpyxl
    wb = openpyxl.load_workbook(path)
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                v = cell.value
                if v is None or isinstance(v, bool):
                    continue
                if isinstance(v, str):
                    if pat.search(v):
                        cell.value = restore_text(v, pat, table)
                elif isinstance(v, (int, float)):
                    s = str(int(v)) if float(v).is_integer() else str(v)
                    if s in table:
                        cell.value = table[s]
                elif isinstance(v, (datetime.date, datetime.datetime)):
                    s = v.strftime("%m/%d/%Y")
                    if s in table:
                        cell.value = table[s]
    wb.save(out)
    hits = []
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str):
                    hits.extend(RE_ANY_ALIAS_SHAPE.findall(cell.value))
    return _residual_issues(sorted(set(hits)))


def _restore_pdf(path, out, pat, table):
    import fitz
    log = []
    doc = fitz.open(path)
    for page in doc:
        words = page_words(page)
        hits = []
        for alias, real in sorted(table.items(), key=lambda p: -len(p[0])):
            for rect in page.search_for(alias):
                r = tuple(rect)
                # Skip only if this hit is fully CONTAINED within an
                # already-claimed (longer) alias's rect — that is a genuine
                # substring occurrence (e.g. 'L001XX' inside the combined
                # name/ID block). An any-overlap test (the earlier version)
                # also fired on adjacent, unrelated lines at normal line
                # spacing, silently leaving those aliases un-restored.
                if any(_rect_contains(h[0], r) for h in hits):
                    continue
                hits.append((r, real))
        if not hits:
            continue
        hit_rects = [h[0] for h in hits]
        room_bounds = [_room_bound(r, words, hit_rects, page.rect.width)
                      for r, _ in hits]
        for (rect, _), x1 in zip(hits, room_bounds):
            x0, y0, _, y1 = rect
            page.add_redact_annot(fitz.Rect(*_inset((x0, y0, x1, y1))))
        try:
            page.apply_redactions(graphics=0)
        except TypeError:
            page.apply_redactions()
        for (rect, real), x1 in zip(hits, room_bounds):
            red = Redaction(rect, real, "restore")
            _draw_alias(page, red, x1, log)
    doc.save(out, garbage=3, deflate=True)
    doc.close()
    text = "\n".join(p.get_text() for p in fitz.open(out))
    issues = [e for e in log if e.startswith("ERROR")]
    if issues:
        issues = ["%d real value(s) could not be redrawn (too long for the "
                  "space the alias occupied) — that student/field is now "
                  "BLANK, not just aliased, in the restored PDF: %s"
                  % (len(issues), issues[:4])]
    issues += _residual_issues(_residual_alias_hits(text))
    return out, issues


# --------------------------------------------------------------------------
# Section 9 — selftest (pure functions, no PDF libraries needed)
# --------------------------------------------------------------------------

def _mk_tokens(text):
    """Build fake word boxes for a course-row string (selftest only)."""
    toks, x = [], 10.0
    for t in text.split():
        w = W(x, 100.0, x + 6.0 * len(t), 108.0, t)
        toks.append(w)
        x = w.x1 + 3.0
    return toks


def selftest():
    ok = True

    def check(cond, label):
        nonlocal ok
        print("  %s  %s" % ("PASS" if cond else "FAIL", label))
        if not cond:
            ok = False

    print("alias spec:")
    a1, a2, a12 = student_aliases(1), student_aliases(133), student_aliases(12)
    check(a1["name"] == "L001XX, F001XX", "student alias shape")
    check(a1["first"][:5] != a2["first"][:5], "unique 5-char first-name prefix")
    check(a1["id"] == "900000001" and RE_OSIS.match(a1["id"]), "synthetic OSIS range")
    check(a1["dob"] == "01/01/1900" and a2["dob"] != a1["dob"], "distinct DOB aliases")
    check(len({student_aliases(i)["dob"] for i in range(1, 500)}) == 499,
          "DOB aliases unique across 499 students")
    check(counselor_alias(3).endswith("CNSL"), "counselor CNSL suffix")
    check(not counselor_alias(1).startswith(("F", "L")), "counselor namespace disjoint")
    check(teacher_alias(5) == "TCH005", "teacher alias shape")
    check(a12["phone"][:9] != a1["id"],
          "phone alias truncated to 9 digits never equals another student's ID")
    check(not any(student_aliases(k)["phone"][:9] == student_aliases(j)["id"]
                  for k in range(1, 200) for j in (k // 10 or 1,)),
          "no phone/id prefix collision across 1..199")

    print("mark classification:")
    for tok, cls in [("85", 1), ("85*", 1), ("100", 1), ("CR*", 1), ("NS*", 1),
                     ("WG*", 1), ("INC", 1), ("B+", 1), ("B*", 1), ("B", 2),
                     ("S", 3), ("E", 2), ("S*", 1), ("NGUYEN", 0), ("M", 0),
                     ("ALG", 0), ("92**", 1), ("HIST", 0)]:
        check(_mark_class(tok) == cls, "mark class %r -> %d" % (tok, cls))

    print("teacher segmentation (handoff traps):")
    cases = [
        # (row, expected_teacher, expected_mark, permissive)
        ("2024 / 1 435 PES41 PHYS ED 85 NGUYEN 1.00/1.00", "NGUYEN", "85", False),
        ("2024 / 1 435 MES41 ALG 2 85 SMITH 1.00/1.00", "SMITH", "85", False),
        ("2024 / 1 435 HES41 US HIST 85 SMITH S 1.00/1.00", "SMITH S", "85", False),
        ("2024 / 2 435 AES41 STUDIO ART 92** O'BRIEN M 1.00/1.00", "O'BRIEN M", "92**", False),
        ("2023 / 1 400 TRF01 TRANSFER PHYS E CR* 1.00/1.00", "", "CR*", False),
        ("2024 / 1 435 MUS41 ORCHESTRA S 1.00/1.00", "", "S", False),
        ("2024 / 1 435 DRA41 DRAMA B QUINCEY 1.00/1.00", "QUINCEY", "B", False),
        ("2024 / 1 435 SCI41 CHEM LAB NS* 0.00/0.00", "", "NS*", False),
        ("2024 / 1 435 ENG41 ENG 4 90 VARGA-QUINN 1.00/1.00", "VARGA-QUINN", "90", False),
        ("2024 / 1 435 HIS41 GLOBAL HIST 3 88 DE LA CRUZ 1.00/1.00", "DE LA CRUZ", "88", False),
        ("2024 / 1 435 SES41 CHEM 92 MUÑOZ 1.00/1.00", "MUÑOZ", "92", False),
        ("2024 / 1 435 SPS41 SPANISH 88 PEÑA 1.00/1.00", "PEÑA", "88", False),
    ]
    for row, t_exp, m_exp, perm in cases:
        seg = segment_course_row(_mk_tokens(row))
        t = " ".join(w.text for w in seg["teacher"]) if seg else "<none>"
        m = (seg["mark"].text if seg and seg["mark"] else "") if seg else "<none>"
        check(seg is not None and t == t_exp and m == m_exp
              and seg["permissive"] == perm,
              "%r -> teacher %r mark %r" % (row[16:], t, m))
    seg = segment_course_row(_mk_tokens("2024 / 1 435 XXX41 MYSTERY ROW QUIMBLETON 1.00/1.00"))
    check(seg is not None and seg["permissive"]
          and "QUIMBLETON" in [w.text for w in seg["teacher"]],
          "anchor-less row takes the permissive path")
    check(segment_course_row(_mk_tokens("Subtotal : 4.00/4.00")) is None,
          "subtotal line is not a course row")
    check(segment_course_row(_mk_tokens("2024 Term 1 SXRK ALGEBRA REG 85")) is None,
          "exam line is not a course row")

    print("column-split merge-back:")
    left = _mk_tokens("2024 / 1 435 EES87 ADVANCED PLACEMENT ENGLISH LIT")
    for w in left:
        pass
    teacher_word = W(295.0, 100.0, 330.0, 108.0, "STRADDLETON")
    credit_word = W(340.0, 100.0, 372.0, 108.0, "1.00/1.00")
    band_words = left + [teacher_word, credit_word]
    rows = logical_rows(band_words)
    merged = [r for r in rows if any(w.text == "STRADDLETON" for w in r)]
    check(len(merged) == 1 and any(RE_CREDITS.match(w.text) for w in merged[0]),
          "row whose credits token crosses x=304 is merged back into one row")
    seg2 = segment_course_row(merged[0]) if merged else None
    check(seg2 is not None and "STRADDLETON" in [w.text for w in seg2["teacher"]],
          "merged cross-boundary row still yields the teacher")

    print("pattern-class scanner:")
    check(RE_SCAN_NAME.search("HEADER TESTLAST, TESTFIRST TRAILER"), "LAST, FIRST hit")
    check(not RE_SCAN_NAME.search("L001XX, F001XX"), "alias name not flagged")
    check(RE_SCAN_NAME.search("NG, KEVIN"), "2-char surname hit")
    check(RE_SCAN_NAME.search("RIVERA, P"), "single-initial given-name hit")
    check(RE_SCAN_NAME.search("PEÑA, JOSÉ"), "unicode name hit")
    check(RE_SCAN_NAME_TITLE.search("see Ramona De La Cruz about"), "Title-Case First-Last hit")
    check(not RE_SCAN_NAME_TITLE.search(_scrub_builtin("Student Permanent Record")),
          "builtin boilerplate scrubbed from Title-Case scan")
    check(not RE_SCAN_NAME_TITLE.search("Page 1 of 2"), "footer not flagged")
    check(RE_SCAN_STREET.search("88 OCEAN VIEW AVE APT"), "street shape hit")
    check(not RE_SCAN_STREET.search("2024 / 1 435 US1 US HIST 3"), "course row not street")
    check(RE_SCAN_ID9.search("212345678"), "9-digit hit")
    check(RE_SCAN_DATE.search("05/12/2008"), "DOB-shaped date hit")

    print("restore direction:")
    m = Mapping(os.devnull + ".nonexistent.json")
    m.data["students"]["212345678"] = {
        "idx": 1, "name": "TESTLAST, TESTFIRST",
        "real": {"dob": "05/12/2008", "phone": "7185550001",
                 "address": "88 FAKE ST NY 11200", "parent": "TESTLAST, PARENTA"}}
    m.data["students"]["212345679"] = {
        "idx": 2, "name": "SECONDLAST, SECONDFIRST",
        "real": {"parent": "Ramona De La Cruz"}}
    m.data["teachers"]["QUINCEY"] = "TCH001"
    m.data["teachers"]["OBRIENZ M"] = "TCH002 M"
    pat, table = _build_restorer(m)
    src = ("L001XX, F001XX / 900000001 dob 01/01/1900 ph 8000000001 "
           "addr ADDR001X par PL001XX, PF001XX t TCH001 t2 TCH002 M solo F001XX "
           "num 1900000001")
    out = restore_text(src, pat, table)
    check("TESTLAST, TESTFIRST / 212345678" in out, "name+id restored")
    check("05/12/2008" in out and "7185550001" in out, "dob+phone restored")
    check("88 FAKE ST NY 11200" in out, "address restored")
    check("TESTLAST, PARENTA" in out, "parent restored")
    check("QUINCEY" in out, "teacher restored")
    check("OBRIENZ M" in out, "teacher alias with initial restored")
    check(out.endswith("num 1900000001"),
          "digit alias embedded in a longer number is NOT corrupted (\\b anchors)")
    check(out.split()[3] == "TESTFIRST" or "TESTFIRST" in out.split(),
          "bare first-name alias restored")
    check("XX" not in out and "TCH0" not in out.replace("num 1900000001", ""),
          "zero residual aliases")

    src2 = "please call PF002XX about attendance, ref PL002XX"
    out2 = restore_text(src2, pat, table)
    check("please call De La Cruz about attendance" in out2
          or "please call Ramona De La" in out2,
          "bare parent given-name alias restores to first-name portion only")
    check(out2.count(",") == src2.count(","),
          "bare parent alias does not introduce a new comma")

    print("\nselftest: %s" % ("PASS" if ok else "FAIL"))
    return ok


# --------------------------------------------------------------------------
# Section 10 — CLI + GUI
# --------------------------------------------------------------------------

def _print_report(report):
    print("\n".join(report))


def run_cli(argv):
    ap = argparse.ArgumentParser(description=__doc__.splitlines()[1])
    ap.add_argument("input", nargs="?", help="record PDF to anonymize")
    ap.add_argument("--restore", metavar="FILE",
                    help="restore real values into a returned file")
    ap.add_argument("--expect", type=int, metavar="N",
                    help="expected unique student count (hard check)")
    ap.add_argument("--mapping", metavar="PATH",
                    help="mapping JSON (default: beside this script)")
    ap.add_argument("--allow-pattern", action="append", default=[],
                    metavar="EXACT",
                    help="whitelist a legit string for the pattern-class scan "
                         "(repeatable) — must match the FULL flagged string "
                         "exactly, e.g. a school name; a partial/substring "
                         "match would silently exempt a real PII hit that "
                         "merely contains it")
    ap.add_argument("--selftest", action="store_true")
    ap.add_argument("--gui", action="store_true")
    args = ap.parse_args(argv)

    if args.selftest:
        return 0 if selftest() else 1
    if args.restore:
        out, issues = restore_file(args.restore, args.mapping)
        if issues:
            print("restored -> %s" % out)
            print("\nWARNING — review before trusting this file:")
            for issue in issues:
                print("  ! " + issue)
            return 1
        print("restored cleanly -> %s" % out)
        return 0
    if args.input:
        ok, out, report = anonymize_pdf(
            args.input, mapping_path=args.mapping,
            expect_students=args.expect, allow_patterns=args.allow_pattern)
        _print_report(report)
        if ok:
            print("\nPASS — safe to upload: %s" % out)
            return 0
        print("\nFAIL — DO NOT UPLOAD%s" % (" (output kept as %s for "
              "inspection)" % out if out else ""))
        return 1
    return run_gui()


def run_gui():
    """Drag-and-drop window (tkinterdnd2) or plain file-picker fallback."""
    try:
        import tkinter as tk
        from tkinter import filedialog, messagebox, scrolledtext
    except Exception:
        print("tkinter unavailable — pass a PDF path (see --help)")
        return 2

    try:
        from tkinterdnd2 import DND_FILES, TkinterDnD
        root = TkinterDnD.Tk()
        dnd = True
    except Exception:
        root = tk.Tk()
        dnd = False

    root.title("NYC DOE Record Anonymizer v%s" % __version__)
    root.geometry("560x420")
    text = scrolledtext.ScrolledText(root, height=16)

    def log_lines(lines):
        text.insert("end", "\n".join(lines) + "\n")
        text.see("end")

    def handle(path):
        path = path.strip("{}")
        try:
            if path.lower().endswith(".pdf") and "_ANON" not in path:
                ok, out, report = anonymize_pdf(path)
                log_lines(report)
                (messagebox.showinfo if ok else messagebox.showerror)(
                    "Anonymizer", ("PASS — safe to upload:\n%s" % out) if ok
                    else "FAIL — DO NOT UPLOAD. See report.")
            else:
                out, issues = restore_file(path)
                if issues:
                    log_lines(["restored -> %s" % out, "WARNING:"] + issues)
                    messagebox.showwarning(
                        "Anonymizer", "Restored, but review before trusting "
                        "this file — see the log for %d issue(s)."
                        % len(issues))
                else:
                    log_lines(["restored cleanly -> %s" % out])
        except Exception as exc:
            messagebox.showerror("Anonymizer", str(exc))

    label = tk.Label(root, relief="ridge", height=4, text=(
        "Drop a record PDF to anonymize,\nor a returned txt/csv/xlsx/pdf to "
        "restore" if dnd else "Use the buttons below"))
    label.pack(fill="x", padx=10, pady=8)
    if dnd:
        label.drop_target_register(DND_FILES)
        label.dnd_bind("<<Drop>>", lambda e: handle(e.data))

    row = tk.Frame(root)
    tk.Button(row, text="Anonymize PDF…", command=lambda: handle(
        filedialog.askopenfilename(filetypes=[("PDF", "*.pdf")]) or "")
    ).pack(side="left", padx=4)
    tk.Button(row, text="Restore file…", command=lambda: handle(
        filedialog.askopenfilename() or "")).pack(side="left", padx=4)
    row.pack(pady=4)
    text.pack(fill="both", expand=True, padx=10, pady=8)
    root.mainloop()
    return 0


if __name__ == "__main__":
    sys.exit(run_cli(sys.argv[1:]))
